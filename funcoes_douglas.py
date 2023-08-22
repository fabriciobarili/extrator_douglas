import mysql.connector
import openpyxl
from datetime import date


mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="ufmt"
)
mycursor = mydb.cursor()


def insert_noticia_pt1(titulo: str, url: str, data: str, site: str):
    sql = "INSERT INTO `tb_noticias` (`TITULO`, `URL`, `DATA`, `SITE`) VALUES (%s, %s, %s, %s)"
    val = (titulo, url, data, site)
    mycursor.execute(sql, val)
    mydb.commit()
    return "Sucesso"


def getNoticias(site: str):
    sql = f"SELECT `url`, `id` FROM `tb_noticias` where not EXISTS (select 1 from tb_texto where ID_NOTICIA = `tb_noticias`.`ID`) and SITE = '{site}'"
    mycursor.execute(sql)
    myresult = mycursor.fetchall()

    return myresult


def insert_noticia(id: str, texto: str):
    sql = "INSERT INTO `tb_texto` (`ID_NOTICIA`, `TEXTO`) VALUES (%s, %s)"
    val = (id, texto)
    mycursor.execute(sql, val)
    mydb.commit()
    return "Sucesso"

def insert_imagens(id: str, texto: str):
    sql = "INSERT INTO `tb_imagens` (`id_noticia`, `url`) VALUES (%s, %s)"
    val = (id, texto)
    mycursor.execute(sql, val)
    mydb.commit()
    return "Sucesso"

def UpdateData_Noticia(id : str, texto : str):
    sql = "UPDATE `tb_noticias` set  `DATA` = %s WHERE ID = %s"
    val = (texto, id)
    mycursor.execute(sql, val)
    mydb.commit()
    return "Sucesso"

def getNoticiasFolhamax(site: str):
    sql = f"SELECT `url`, `id` FROM `tb_noticias` where not EXISTS (select 1 from tb_texto where ID_NOTICIA = `tb_noticias`.`ID`) and SITE = '{site}'"
    mycursor.execute(sql)
    myresult = mycursor.fetchall()
    return myresult


def ExportExcel():
    nome = input("Digite o nome do seu arquivo sem espaços e sem acentos. Ex:pesquisa_pessoa: ")

    book = openpyxl.Workbook()
    sheet = book.active
    sql = f"SELECT A.ID, A.SITE, A.TITULO, A.DATA, B.TEXTO from tb_noticias as A inner join tb_texto as B on a.ID = b.ID_NOTICIA"
    mycursor.execute(sql)
    myresult = mycursor.fetchall()

    #Montando o cabeçalho
    cell = sheet.cell(row=1, column=1)
    cell.value = "ID"
    cell = sheet.cell(row=1, column=2)
    cell.value = "SITE"
    cell = sheet.cell(row=1, column=3)
    cell.value = "TITULO"
    cell = sheet.cell(row=1, column=4)
    cell.value = "DATA"
    cell = sheet.cell(row=1, column=5)
    cell.value = "TEXTO"


    i = 1
    for row in myresult:
        i += 1
        j = 1
        for col in row:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1



    book.save(f"arquivos/{nome}_{date.today()}.xlsx")

    book = openpyxl.Workbook()
    sheet = book.active
    sql = f"SELECT id_noticia, url from tb_imagens"
    mycursor.execute(sql)
    myresult = mycursor.fetchall()

    cell = sheet.cell(row=1, column=1)
    cell.value = "NOTÍCIA"
    cell = sheet.cell(row=1, column=2)
    cell.value = "URL"

    i = 1
    for row in myresult:
        i += 1
        j = 1
        for col in row:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1

    book.save(f"arquivos/{nome}_{date.today()}_imgs.xlsx")

    return "Sucesso"

def LimparBase():
    PROMPT = input("DESEJA REALMENTE APAGAR TODO O CONTEÚDO DAS TRÊS BASES? S - SIM ; N - NÃO: ").upper()
    if PROMPT == "S":
        sql = "TRUNCATE `tb_imagens`"
        mycursor.execute(sql)

        sql = "TRUNCATE `tb_noticias`"
        mycursor.execute(sql)

        sql = "TRUNCATE `tb_texto`"
        mycursor.execute(sql)


        print("Base limpa")
    else:
        print("Ação cancelada")

    return "Sucesso"